# Decompiled with PyLingual (https://pylingual.io)
# Internal filename: main.py
# Bytecode version: 3.10.0rc2 (3439)
# Source timestamp: 1970-01-01 00:00:00 UTC (0)

import openpyxl
from datetime import datetime
from colorama import init, Fore, Style
import os
import json
import glob
import time
import random
import logging
init(convert=True)
logging.basicConfig(filename='logs.log', level=logging.WARNING)
logging.getLogger('alright').propagate = False
from alright import WhatsApp
logging.getLogger('alright').propagate = False

def check_prerequisite():
    global phone_numbers
    global settings
    excel_files = glob.glob('*.xls*')
    assert len(excel_files) > 0, 'No Excel File Found!'
    assert os.path.isfile('settings.json'), 'settings.json File missing!'
    settings = json.load(open('./settings.json', encoding='utf-8', errors='ignore'))
    for i, file in enumerate(excel_files):
        print(f' |- {i + 1:02}) {file}')
    selected_index = int(input('Enter the number of the file you want to select: ')) - 1
    if selected_index >= 0 and selected_index < len(excel_files):
        excel_file = excel_files[selected_index]
    else:
        raise Exception(f'{Fore.RED}Invalid selection{Fore.RESET}')
    wb = openpyxl.load_workbook(excel_file)
    sheet = wb.worksheets[0]
    for row in sheet.iter_rows(values_only=True):
        headers = row
        break
    for row in sheet['1']:
        if 'number' in str(row.value).lower():
            phone_number_index = headers.index(row.value)
    phone_numbers = [row[phone_number_index] for row in sheet.iter_rows(values_only=True)]
    phone_numbers.pop(0)
    assert len(phone_numbers) != 0, '0 phone number(s) found!'
    phone_numbers = [str(settings['country_code']) + str(num) if len(str(num)) == 10 else str(num) for num in phone_numbers]
    phone_numbers = sorted(set(phone_numbers))
    print(f'{Fore.BLUE}{len(phone_numbers)} phone numbers found{Fore.RESET}')

def main():
    mm = input("Do you want to send multiple messages? ('y' for yes)  >>> ")
    multiple_message = True if mm.lower() == 'y' else False
    i = input("Do you want to send images/docs/videos? ('y' for yes)  >>> ")
    if i.lower() == 'y':
        sv = input("videos? ('y' for yes)  >>> ")
        send_video = True if sv.lower() == 'y' else False
        si = input("images? ('y' for yes)  >>> ")
        send_image = True if si.lower() == 'y' else False
        sd = input("docs? ('y' for yes)  >>> ")
        send_docs = True if sd.lower() == 'y' else False
    else:
        send_docs = send_image = send_video = False
    if send_video:
        sv = str(input('Enter the name of video (and extension)  >>> '))
        assert os.path.isfile(sv), f'{sv} File missing!'
    if send_image:
        si = str(input('Enter the name of image (and extension)  >>> '))
        assert os.path.isfile(si), f'{si} File missing!'
        image_message = str(input('Add text to accompany image (press `Enter` to use message from text file) >>> '))
    if send_docs:
        sd = str(input('Enter the name of doc [a.pdf, b.pdf] for multiple files >>> '))
        sd = [x.strip() for x in sd.split(',')]
        for mud in sd:
            assert os.path.isfile(mud), f'{mud} File missing!'
    if multiple_message:
        message = open('./message.txt', 'r').read().splitlines()
        message = [m.replace('{new_line}', '\n') for m in message]
    else:
        message = open('./message.txt', 'r').read()
        message = message.replace('{new_line}', '\n')
    input(f'{Fore.YELLOW}Message is empty! Press any key to continue...{Fore.RESET}') if message in ('', []) else None
    messenger = WhatsApp()
    for number in phone_numbers:
        try:
            user_status = messenger.find_user(number)
            if not user_status:
                raise Exception(f'Not on whatsapp {number}')
            if send_image:
                messenger.send_picture(si, '') if image_message.strip() == '' else messenger.send_picture(si, image_message)
            if send_video:
                messenger.send_video(sv)
            if send_docs:
                for send_doc in sd:
                    messenger.send_file(send_doc)
                    time.sleep(3)
            if multiple_message:
                for msg in message:
                    messenger.send_message(msg)
            else:
                messenger.send_message(message)
            print(f"{Fore.GREEN}[{datetime.now().strftime('%H:%M:%S')}] Message sent to {number[2:]}{Fore.RESET}")
            time.sleep(random.randint(settings.get('min_wait_time', 0), settings.get('max_wait_time', 2)))
        except Exception as e:
            print(f"{Fore.RED}[{datetime.now().strftime('%H:%M:%S')}] Error - {e} [{number[2:]}]{Fore.RESET}")
    input('Press any key to exit...')
if __name__ == '__main__':
    try:
        check_prerequisite()
        main()
    except AssertionError as e:
        print(f'{Fore.RED}{e}{Fore.RESET}')
    except Exception as e:
        print(f'{Fore.RED}{e}{Fore.RESET}')